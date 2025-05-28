import requests #爬蟲使用模組
import bs4
import openpyxl

res=requests.get("https://www.104.com.tw/jobs/search/?jobsource=index_s&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&mode=s&page=1") #網址是字串
soup = bs4.BeautifulSoup(res.text) 

wb = openpyxl.Workbook() #W要大寫
ws = wb.active #開新視窗
ws["A1"] = "職缺名稱"
ws["B1"] = "職缺連結"
ws["C1"] = "公司名稱"
ws["D1"] = "工作區域"
ws["E1"] = "薪資待遇"
ws["F1"] = "計薪方式"
ws["G1"] = "薪資(起薪)"
ws["H1"] = "薪資(最高)"
ws["I1"] = "薪資(平均)"
ws["J1"] = "工作縣市"
ws["K1"] = "鄉鎮市區"


page=1
while soup.find_all("div",class_="info-container") !=[]: #![] 頁面跑到沒有list資料為止
    print("----------------------")
    print("第" , page, "頁")
    print("----------------------")
    
    for job in soup.find_all("div",class_="info-container"):
        a = job.a.text #職缺名稱 
        b = job.a["href"] #職缺連結
        c = job.select("div")[1].text
        d = job.select("div")[2].select("span")[0].text #工作地點
        #拆解工作縣市、鄉鎮
        j = d[:3]
        k = d[3:]
        
        e = job.select("div")[2].select("span")[3].text #薪資待遇
        #取計薪方式
        f = e[:2]
        if f=="待遇":
            f = "面議"
        
        if f=="論件":
            f = "論件計酬"
        
        #薪水
        salary = ""
        for char in e:
            if char=="0" or char=="1" or char=="2" or char=="3" or char=="4" or char=="5" or char=="6" or char=="7" or char=="8" or char=="9" or char=="~" :
                salary+=char
        
        #拆解薪支上下限   
        if "~" in salary:
            low = salary[:salary.find("~")]
            high = salary[salary.find("~")+1:]
        else:
            low = salary
            high = salary
        #計算平均薪資
        if low != "" and high != "":
            low = int(low)
            high = int(high)
            avg = (low + high)/2
        else:
            avg = ""

        f = f
        g = low
        h = high
        i = avg

        ws.append([a,b,c,d,e,f,g,h,i,j,k])
    
    page+=1
    res=requests.get("https://www.104.com.tw/jobs/search/?jobsource=index_s&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&mode=s&page="+str(page)) #正常而言網址為對應的href，104要手動調整資料格式
    soup = bs4.BeautifulSoup(res.text)
    

    wb.save("./output/104找工作_大數據分析_0124.xlsx")